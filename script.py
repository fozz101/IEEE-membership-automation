# -*- coding: utf-8 -*-
"""
Created on Fri Jan  7 14:22:09 2023

@author: fedig
"""

from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support.select import Select
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import time
import pandas as pd
import numpy as np
import openpyxl
from openpyxl.styles import PatternFill
import string
import tkinter as tk


def progressBarLength(excelFile):
    members = pd.read_excel(excelFile)
    compt=0
    try:
        for i in members["Status"]:
            if i!="Done":
                compt+=1
        return compt
    except KeyError:
        tk.messagebox.showerror('Error', "'Status' Column doesn't EXIST !!")


def getStatusIndex(excelFile):
    letters = list(string.ascii_uppercase)
    members = pd.read_excel(excelFile)
    compt=0
    for i in members.columns:
        compt+=1
    
    return letters[compt-1]

def addStatusColumn(excelFile):
    members = pd.read_excel(excelFile)
    flag = False
    for i in members.columns:
        if i.upper() =="STATUS":
            return 0
    members["Status"]= "Undone"
    members.to_excel(excelFile)


def splitExcel(excelFile,nbreOfExcelFiles):
    addStatusColumn(excelFile)
    members = pd.read_excel(excelFile)
    split = np.array_split(members, nbreOfExcelFiles)
    fileName = excelFile.split("/")[len(excelFile.split("/"))-1]
    
    path =""
    
    for i in range (len(excelFile.split("/"))-2):
        path+=excelFile.split("/")[i]
    
    
    for i in range(len(split)):
        split[i].to_excel(fileName[:len(fileName)-4]+str(i)+'.xlsx')
        #Delete 1st column
        '''
        book = openpyxl.load_workbook(path+fileName[:len(fileName)-4]+str(i)+'.xlsx')
        sheet = book.active
        sheet.delete_cols(1)
        book.save(path+fileName[:len(fileName)-4]+str(i)+'.xlsx')
        '''
        


def checkMembership(membership,memberships,driver,url):
    if membership in memberships:
        driver.get(url)
        try:
            addItemBtn = driver.find_element(By.ID, "addItems")
            addItemBtn.click()
        except Exception :
            pass

        
def proceedToPayment(driver,cardNumber,cardYear,cardMonth,cardOwner):
    driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
    continueBtnToCartPayment = driver.find_element(By.XPATH, "//div[6]/input")
    #time.sleep(10)
    continueBtnToCartPayment.click()
    #time.sleep(10)
    continueBtnToCartPayment1 = driver.find_element(By.ID, "proceed-to-checkout-btn-sec")

    driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
    continueBtnToCartPayment1.click()
    
    
    creditCardRadio = driver.find_element(By.ID, "op-ccpay")
    creditCardRadio.click()
    
    billingAddressRadio = driver.find_element(By.ID, "sameas_mailing_address")
    billingAddressRadio.click()
    #time.sleep(5)
    
    creditCardTypeSelectElement = driver.find_element(By.ID, "choose-card-type")
    creditCardSelect = Select(creditCardTypeSelectElement)
    creditCardSelect.select_by_visible_text("MasterCard")
    
    creditCardNumber = driver.find_element(By.ID,"credit-card-number")
    creditCardNumber.send_keys(cardNumber)
    
    creditCardYearSelectElement = driver.find_element(By.ID, "expiration-year")
    creditCardYearSelect = Select(creditCardYearSelectElement)
    creditCardYearSelect.select_by_visible_text(cardYear)
 
    
    creditCardMonthSelectElement = driver.find_element(By.ID, "month")
    creditCardMonthSelect = Select(creditCardMonthSelectElement)
    creditCardMonthSelect.select_by_visible_text(cardMonth)
    
    nameOncard = driver.find_element(By.ID,"name-on-card")
    nameOncard.send_keys(cardOwner)
    time.sleep(7)
    verifyCardBtn = driver.find_element(By.XPATH, "//div[10]/div/input")
    verifyCardBtn.click()
    time.sleep(10)

    continuePaymentBtn = driver.find_element(By.XPATH, "//div[14]/div/input")
    continuePaymentBtn.click()    
    termsConditionsCheckbox1 = driver.find_element(By.ID, "terms-conditions")
    termsConditionsCheckbox2 = driver.find_element(By.ID, "membership-terms-conditions")
    
    termsConditionsCheckbox1.click()
    termsConditionsCheckbox2.click() 
    
    closeTermsBtn = driver.find_element(By.ID, "ibpModalClose")
    closeTermsBtn.click()
    
    ####finishPaymentBtn = driver.find_element(By.ID, "terms-conditions")
    
    
    
def setUpAccountNoPayment(email,password,memberships,memberName,memberID):
    #os.chdir("D:/IEEEMembershipAutomation/")
    service = Service(executable_path="/chromedriver")
    driver = webdriver.Chrome(service=service)
    driver.maximize_window()
    driver.implicitly_wait(10)
    driver.execute_script('document.getElementsByTagName("html")[0].style.scrollBehavior = "auto"')
    
    
    driver.get("https://www.ieee.org/")
    
    
    cookiesBtn = driver.find_element(By.LINK_TEXT, "Accept & Close")
    signInLink = driver.find_element(By.LINK_TEXT, "Sign In")
    cookiesBtn.click()
    signInLink.click()
    
    emailTextbox = driver.find_element(by=By.NAME, value="pf.username")
    passwordTextbox = driver.find_element(by=By.NAME, value="pf.pass")
    signInBtn = driver.find_element(By.ID, "modalWindowRegisterSignInBtn")
    
    emailTextbox.send_keys(email)
    passwordTextbox.send_keys(password)
    signInBtn.click()
    
    driver.get("https://www.ieee.org/membership-application/join.html?grade=Student")
    
    
    
    
    driver.find_element(By.NAME, "customer.addresses[0].line1").clear()
    driver.find_element(By.NAME, "customer.addresses[0].city").clear()
    driver.find_element(By.NAME, "customer.addresses[0].postalCode").clear()
    
    
    '''
    countrySelectElement = driver.find_element(By.ID, "country")
    countrySelect = Select(countrySelectElement)
    countrySelect.select_by_visible_text('Tunisia')
    '''
    addressTextbox = driver.find_element(By.NAME, "customer.addresses[0].line1")
    cityTextbox = driver.find_element(By.NAME, "customer.addresses[0].city")
    zipTextbox = driver.find_element(By.NAME, "customer.addresses[0].postalCode")
    continueBtn = driver.find_element(by=By.XPATH, value= "//div/div[6]/input")
    addressType = driver.find_element(By.XPATH, "//div[3]/input")
    provinceSelectElement = driver.find_element(By.ID, "province")
    provinceSelect = Select(provinceSelectElement)
    
    cartContent = driver.find_element(By.CLASS_NAME, "mc-section").text
    #print(cartContent)
    if addressType.is_selected():
        pass
    else:    
        addressType.click()
    
    addressTextbox.send_keys("ESPRIT, Pole Technologique - El Ghazala")
    cityTextbox.send_keys("El Ghazala")
    zipTextbox.send_keys("2083")
    provinceSelect.select_by_visible_text('Ariana')
    
    continueBtn.click()
    

    
    #enajem nsala7ha b try catch student w professional ken catcha error ibadel
    try:
        undergradRadio = driver.find_element(By.ID, "studentStatusUndergraduate Student")
    except Exception:
        time.sleep(7)
        studentSelectElement = WebDriverWait(driver, 30).until(EC.presence_of_element_located((By.NAME, "careerPhase")))
        studentSelect = Select(studentSelectElement)
        studentSelect.select_by_index(1)
        time.sleep(5)
        undergradRadio = driver.find_element(By.ID, "studentStatusUndergraduate Student")
    
    
    
    
    
    directoryCheckbox = driver.find_element(By.ID, "memberdir-options1")
    whyJoin1Checkbox = driver.find_element(By.ID, "TechnicallyCurrent")
    whyJoin2Checkbox = driver.find_element(By.ID, "CareerOpurtunities")
    whyJoin3Checkbox = driver.find_element(By.ID, "ExpandProfessionalNetwork")
    whyJoin4Checkbox = driver.find_element(By.ID, "ConnectToLocalActivities")
    continue2Btn = driver.find_element(By.XPATH, "//div[5]/div/div[2]/div/input")
    
    
    
    stateSelectElement = driver.find_element(By.ID, "state")
    stateSelect = Select(stateSelectElement)
    stateSelect.select_by_visible_text('Ariana')
    
    
    
    academicSelectElement = driver.find_element(By.ID, "stud-academic-program")
    academicSelect = Select(academicSelectElement)
    academicSelect.select_by_visible_text("Computer Science")
    
    gradMSelectElement = driver.find_element(By.ID, "estimated-grad-month")
    gradMSelect = Select(gradMSelectElement)
    gradMSelect.select_by_visible_text('June')
    
    gradYSelectElement = driver.find_element(By.ID, "estimated-grad-year")
    gradYSelect = Select(gradYSelectElement)
    gradYSelect.select_by_visible_text('2028')
    
    studySelectElement = driver.find_element(By.ID, "stud-current-study")
    studySelect = Select(studySelectElement)
    studySelect.select_by_visible_text("Computer Sciences and Information Technologies")
    
    
    
    
    
    techSelectElement = driver.find_element(By.ID, "stud-technical-focus")
    techSelect = Select(techSelectElement)
    techSelect.select_by_visible_text("Computing and Processing (Hardware/Software)")
    
    
    
    
    uniSearchBtn = driver.find_element(By.ID, "searchUniversity")
    uniSearchBtn.click()
    
    uniFilterTextbox = driver.find_element(By.ID, "universityFilter")
    uniFilterTextbox.send_keys("esprit")
    
    espritLinktext = driver.find_element(By.LINK_TEXT, "Private Higher School of Engineering and Technology (ESPRIT) (ESPRIT)")
    espritLinktext.click()
    undergradRadio.click()
    directoryCheckbox.click()
    
    degreeSelectElement = driver.find_element(By.ID, "stud-degree-pursued")
    degreeSelect = Select(degreeSelectElement)
    degreeSelect.select_by_visible_text("Engineer")
    
    
    if whyJoin1Checkbox.is_selected():
        pass
    else:    
        whyJoin1Checkbox.click()
        
    if whyJoin2Checkbox.is_selected():
        pass
    else:    
        whyJoin2Checkbox.click()
        
    
    if whyJoin3Checkbox.is_selected():
        pass
    else:    
        whyJoin3Checkbox.click()
        
    
    if whyJoin4Checkbox.is_selected():
        pass
    else:    
        whyJoin4Checkbox.click()
        
    
    accSelectElement = driver.find_element(By.ID, "stud-prog-accredited")
    accSelect = Select(accSelectElement)
    accSelect.select_by_value("Yes")
    
    driver.find_element(By.ID, "referring-mem-name").clear()
    driver.find_element(By.ID, "referring-mem-number").clear()
    referringName = driver.find_element(By.ID, "referring-mem-name")
    referringID = driver.find_element(By.ID, "referring-mem-number")
    
    referringName.send_keys(memberName)
    referringID.send_keys(memberID)
    
    #96246048
    
    
    
    continue2Btn.click()
    time.sleep(1)
    
    if "Computer" not in cartContent:
        checkMembership("CS", memberships, driver, "https://www.ieee.org/membership-catalog/productdetail/showProductDetailPage.html?product=MEMC016&searchResults=Y")
    if "Computational" not in cartContent:
        checkMembership("CIS", memberships, driver, "https://www.ieee.org/membership-catalog/productdetail/showProductDetailPage.html?product=MEMCIS011&searchResults=Y")
    if "Power" not in cartContent:
        checkMembership("PES", memberships, driver, "https://www.ieee.org/membership-catalog/productdetail/showProductDetailPage.html?product=MEMPE031&searchResults=Y")
    if "Industry" not in cartContent:    
        checkMembership("IAS", memberships, driver, "https://www.ieee.org/membership-catalog/productdetail/showProductDetailPage.html?product=MEMIA034&searchResults=Y")
    if "Microwave" not in cartContent:    
        checkMembership("MTTS", memberships, driver, "https://www.ieee.org/membership-catalog/productdetail/showProductDetailPage.html?product=MEMMTT017&searchResults=Y")
    if "Industrial" not in cartContent:    
        checkMembership("IES", memberships, driver, "https://www.ieee.org/membership-catalog/productdetail/showProductDetailPage.html?product=MEMIE013&searchResults=Y")
    if "Robotics" not in cartContent:    
        checkMembership("RAS", memberships, driver, "https://www.ieee.org/membership-catalog/productdetail/showProductDetailPage.html?product=MEMRA024&searchResults=Y")
    if "SIGHT" not in cartContent:    
        checkMembership("SIGHT", memberships, driver, "https://www.ieee.org/membership-catalog/productdetail/showProductDetailPage.html?product=MEMSIGHT&searchResults=Y")
    if "Women" not in cartContent:    
        checkMembership("WIE", memberships, driver, "https://www.ieee.org/membership-catalog/productdetail/showProductDetailPage.html?product=MEMWIE050&searchResults=Y")
    if "Aerospace" not in cartContent:    
        checkMembership("AESS", memberships, driver, "https://www.ieee.org/membership-catalog/productdetail/showProductDetailPage.html?product=MEMAES010&searchResults=Y")
        
    time.sleep(2)
    '''
    driver.get("https://www.ieee.org/membership-catalog/index.html?N=0")
    time.sleep(2)
    '''
    #proceed to payment
    #proceedToPayment(driver,cardNumber,cardYear,cardMonth,cardOwner)

    

    driver.quit()
def setUpAccountWithPayment(email,password,memberships,memberName,memberID,cardNumber,cardYear,cardMonth,cardOwner):
    #os.chdir("D:/IEEEMembershipAutomation/")
    service = Service(executable_path="/chromedriver")
    driver = webdriver.Chrome(service=service)
    driver.maximize_window()
    driver.implicitly_wait(10)
    driver.execute_script('document.getElementsByTagName("html")[0].style.scrollBehavior = "auto"')
    
    
    driver.get("https://www.ieee.org/")
    
    
    cookiesBtn = driver.find_element(By.LINK_TEXT, "Accept & Close")
    signInLink = driver.find_element(By.LINK_TEXT, "Sign In")
    cookiesBtn.click()
    signInLink.click()
    
    emailTextbox = driver.find_element(by=By.NAME, value="pf.username")
    passwordTextbox = driver.find_element(by=By.NAME, value="pf.pass")
    signInBtn = driver.find_element(By.ID, "modalWindowRegisterSignInBtn")
    
    emailTextbox.send_keys(email)
    passwordTextbox.send_keys(password)
    signInBtn.click()
    
    driver.get("https://www.ieee.org/membership-application/join.html?grade=Student")
    
    
    
    
    driver.find_element(By.NAME, "customer.addresses[0].line1").clear()
    driver.find_element(By.NAME, "customer.addresses[0].city").clear()
    driver.find_element(By.NAME, "customer.addresses[0].postalCode").clear()
    
    
    '''
    countrySelectElement = driver.find_element(By.ID, "country")
    countrySelect = Select(countrySelectElement)
    countrySelect.select_by_visible_text('Tunisia')
    '''
    addressTextbox = driver.find_element(By.NAME, "customer.addresses[0].line1")
    cityTextbox = driver.find_element(By.NAME, "customer.addresses[0].city")
    zipTextbox = driver.find_element(By.NAME, "customer.addresses[0].postalCode")
    continueBtn = driver.find_element(by=By.XPATH, value= "//div/div[6]/input")
    addressType = driver.find_element(By.XPATH, "//div[3]/input")
    provinceSelectElement = driver.find_element(By.ID, "province")
    provinceSelect = Select(provinceSelectElement)
    
    cartContent = driver.find_element(By.CLASS_NAME, "mc-section").text
    #print(cartContent)
    if addressType.is_selected():
        pass
    else:    
        addressType.click()
    
    addressTextbox.send_keys("ESPRIT, Pole Technologique - El Ghazala")
    cityTextbox.send_keys("El Ghazala")
    zipTextbox.send_keys("2083")
    provinceSelect.select_by_visible_text('Ariana')
    
    continueBtn.click()
    

    
    #enajem nsala7ha b try catch student w professional ken catcha error ibadel
    try:
        undergradRadio = driver.find_element(By.ID, "studentStatusUndergraduate Student")
    except Exception:
        time.sleep(7)
        studentSelectElement = WebDriverWait(driver, 30).until(EC.presence_of_element_located((By.NAME, "careerPhase")))
        studentSelect = Select(studentSelectElement)
        studentSelect.select_by_index(1)
        time.sleep(5)
        undergradRadio = driver.find_element(By.ID, "studentStatusUndergraduate Student")
    
    
    
    
    
    directoryCheckbox = driver.find_element(By.ID, "memberdir-options1")
    whyJoin1Checkbox = driver.find_element(By.ID, "TechnicallyCurrent")
    whyJoin2Checkbox = driver.find_element(By.ID, "CareerOpurtunities")
    whyJoin3Checkbox = driver.find_element(By.ID, "ExpandProfessionalNetwork")
    whyJoin4Checkbox = driver.find_element(By.ID, "ConnectToLocalActivities")
    continue2Btn = driver.find_element(By.XPATH, "//div[5]/div/div[2]/div/input")
    
    
    
    stateSelectElement = driver.find_element(By.ID, "state")
    stateSelect = Select(stateSelectElement)
    stateSelect.select_by_visible_text('Ariana')
    
    
    
    academicSelectElement = driver.find_element(By.ID, "stud-academic-program")
    academicSelect = Select(academicSelectElement)
    academicSelect.select_by_visible_text("Computer Science")
    
    gradMSelectElement = driver.find_element(By.ID, "estimated-grad-month")
    gradMSelect = Select(gradMSelectElement)
    gradMSelect.select_by_visible_text('June')
    
    gradYSelectElement = driver.find_element(By.ID, "estimated-grad-year")
    gradYSelect = Select(gradYSelectElement)
    gradYSelect.select_by_visible_text('2028')
    
    studySelectElement = driver.find_element(By.ID, "stud-current-study")
    studySelect = Select(studySelectElement)
    studySelect.select_by_visible_text("Computer Sciences and Information Technologies")
    
    
    
    
    
    techSelectElement = driver.find_element(By.ID, "stud-technical-focus")
    techSelect = Select(techSelectElement)
    techSelect.select_by_visible_text("Computing and Processing (Hardware/Software)")
    
    
    
    
    uniSearchBtn = driver.find_element(By.ID, "searchUniversity")
    uniSearchBtn.click()
    
    uniFilterTextbox = driver.find_element(By.ID, "universityFilter")
    uniFilterTextbox.send_keys("esprit")
    
    espritLinktext = driver.find_element(By.LINK_TEXT, "Private Higher School of Engineering and Technology (ESPRIT) (ESPRIT)")
    espritLinktext.click()
    undergradRadio.click()
    directoryCheckbox.click()
    
    degreeSelectElement = driver.find_element(By.ID, "stud-degree-pursued")
    degreeSelect = Select(degreeSelectElement)
    degreeSelect.select_by_visible_text("Engineer")
    
    
    if whyJoin1Checkbox.is_selected():
        pass
    else:    
        whyJoin1Checkbox.click()
        
    if whyJoin2Checkbox.is_selected():
        pass
    else:    
        whyJoin2Checkbox.click()
        
    
    if whyJoin3Checkbox.is_selected():
        pass
    else:    
        whyJoin3Checkbox.click()
        
    
    if whyJoin4Checkbox.is_selected():
        pass
    else:    
        whyJoin4Checkbox.click()
        
    
    accSelectElement = driver.find_element(By.ID, "stud-prog-accredited")
    accSelect = Select(accSelectElement)
    accSelect.select_by_value("Yes")
    
    driver.find_element(By.ID, "referring-mem-name").clear()
    driver.find_element(By.ID, "referring-mem-number").clear()
    referringName = driver.find_element(By.ID, "referring-mem-name")
    referringID = driver.find_element(By.ID, "referring-mem-number")
    
    referringName.send_keys(memberName)
    referringID.send_keys(memberID)
    
    #96246048
    
    
    
    continue2Btn.click()
    time.sleep(1)
    
    if "Computer" not in cartContent:
        checkMembership("CS", memberships, driver, "https://www.ieee.org/membership-catalog/productdetail/showProductDetailPage.html?product=MEMC016&searchResults=Y")
    if "Computational" not in cartContent:
        checkMembership("CIS", memberships, driver, "https://www.ieee.org/membership-catalog/productdetail/showProductDetailPage.html?product=MEMCIS011&searchResults=Y")
    if "Power" not in cartContent:
        checkMembership("PES", memberships, driver, "https://www.ieee.org/membership-catalog/productdetail/showProductDetailPage.html?product=MEMPE031&searchResults=Y")
    if "Industry" not in cartContent:    
        checkMembership("IAS", memberships, driver, "https://www.ieee.org/membership-catalog/productdetail/showProductDetailPage.html?product=MEMIA034&searchResults=Y")
    if "Microwave" not in cartContent:    
        checkMembership("MTTS", memberships, driver, "https://www.ieee.org/membership-catalog/productdetail/showProductDetailPage.html?product=MEMMTT017&searchResults=Y")
    if "Industrial" not in cartContent:    
        checkMembership("IES", memberships, driver, "https://www.ieee.org/membership-catalog/productdetail/showProductDetailPage.html?product=MEMIE013&searchResults=Y")
    if "Robotics" not in cartContent:    
        checkMembership("RAS", memberships, driver, "https://www.ieee.org/membership-catalog/productdetail/showProductDetailPage.html?product=MEMRA024&searchResults=Y")
    if "SIGHT" not in cartContent:    
        checkMembership("SIGHT", memberships, driver, "https://www.ieee.org/membership-catalog/productdetail/showProductDetailPage.html?product=MEMSIGHT&searchResults=Y")
    if "Women" not in cartContent:    
        checkMembership("WIE", memberships, driver, "https://www.ieee.org/membership-catalog/productdetail/showProductDetailPage.html?product=MEMWIE050&searchResults=Y")
    if "Aerospace" not in cartContent:    
        checkMembership("AESS", memberships, driver, "https://www.ieee.org/membership-catalog/productdetail/showProductDetailPage.html?product=MEMAES010&searchResults=Y")
        
    time.sleep(2)
    '''
    driver.get("https://www.ieee.org/membership-catalog/index.html?N=0")
    time.sleep(2)
    '''
    #proceed to payment
    proceedToPayment(driver,cardNumber,cardYear,cardMonth,cardOwner)

    

    driver.quit()



def mainNoPayment(excelFile,memberName,memberID):
    nbrOfUndoneAccounts = progressBarLength(excelFile)
    addStatusColumn(excelFile)
    colLetter=getStatusIndex(excelFile)
    members = pd.read_excel(excelFile)
    wb = openpyxl.load_workbook(excelFile)
    sheet = wb.active    
    green = "007FFFD4"
    red = "00C70039"
    '''
    setUpAccount(members["EmailAddress"][0],members["Password"][0],members["Memberships"][0])
    
    
    for ind in members.index:
        if members["Status"][ind]!="Done":
            setUpAccount(members["EmailAddress"][ind],members["Password"][ind],members["Memberships"][ind])
    '''
    
    for ind in members.index:
        try:
            if members["Status"][ind]!="Done":
                setUpAccountNoPayment(members["EmailAddress"][ind],members["Password"][ind],members["Memberships"][ind],memberName,memberID)
                sheet[colLetter+str(ind+2)].value="Done"
                sheet[colLetter+str(ind+2)].fill = PatternFill(start_color=green, end_color=green,fill_type = "solid")
                wb.save(excelFile)
                
                #progressBar["value"] +=int(100/nbrOfUndoneAccounts)
                
            else:
                pass
        except Exception :
            sheet[colLetter+str(ind+2)].value="Undone"
            sheet[colLetter+str(ind+2)].fill = PatternFill(start_color=red, end_color=red,fill_type = "solid")
            wb.save(excelFile)
    
    tk.messagebox.showinfo("Done", "- Done !")
    print("FINISHED !")
    

def mainWithPayment(excelFile,memberName,memberID,cardNumber,cardYear,cardMonth,cardOwner):
    
    addStatusColumn(excelFile)
    colLetter=getStatusIndex(excelFile)
    members = pd.read_excel(excelFile)
    wb = openpyxl.load_workbook(excelFile)
    sheet = wb.active    
    green = "007FFFD4"
    red = "00C70039"
    
    #setUpAccount(members["EmailAddress"][0],members["Password"][0],members["Memberships"][0],memberName,memberID,cardNumber,cardYear,cardMonth,cardOwner)
    
    '''
    for ind in members.index:
        if members["Status"][ind]!="Done":
            setUpAccountWithPayment(members["EmailAddress"][ind],members["Password"][ind],members["Memberships"][ind],memberName,memberID,cardNumber,cardYear,cardMonth,cardOwner)
    '''
    
    for ind in members.index:
        try:
            if members["Status"][ind]!="Done":
                setUpAccountWithPayment(members["EmailAddress"][ind],members["Password"][ind],members["Memberships"][ind],memberName,memberID,cardNumber,cardYear,cardMonth,cardOwner)
                sheet[colLetter+str(ind+2)].value="Done"
                sheet[colLetter+str(ind+2)].fill = PatternFill(start_color=green, end_color=green,fill_type = "solid")
                wb.save(excelFile)
            else:
                pass
        except Exception :
            sheet[colLetter+str(ind+2)].value="Undone"
            sheet[colLetter+str(ind+2)].fill = PatternFill(start_color=red, end_color=red,fill_type = "solid")
            wb.save(excelFile)
     
    tk.messagebox.showinfo("Done", "- Done !")
    print("FINISHED !")



#mainWithPayment("[Automation] IEEE Membership (Responses).xlsx","memberName","96246048","5359563923720646","2030","June","Salem")
